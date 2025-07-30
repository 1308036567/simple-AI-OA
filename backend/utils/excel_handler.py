import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Optional, Tuple
import io
import logging
from datetime import datetime, date
import re
from config import EMPLOYEE_FIELD_MAPPING
from utils.validation import DataValidator, validate_chinese_name, validate_id_card, validate_phone, validate_bank_card

logger = logging.getLogger(__name__)

class ExcelHandler:
    """Excel文件处理器"""
    
    def __init__(self):
        self.field_mapping = EMPLOYEE_FIELD_MAPPING
    
    def read_employee_excel(self, file_content: bytes) -> Dict[str, Any]:
        """
        读取员工Excel文件
        
        Args:
            file_content: Excel文件内容
            
        Returns:
            包含员工数据和错误信息的字典
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
            
            # 清理数据
            df = df.dropna(how='all')  # 删除空行
            df = df.fillna('')  # 填充空值
            
            # 标准化列名
            df.columns = [str(col).strip() for col in df.columns]
            
            # 映射字段
            mapped_data = []
            errors = []
            
            for index, row in df.iterrows():
                employee_data = self._map_employee_fields(row.to_dict(), index + 1)
                
                # 验证数据
                validation_errors = DataValidator.validate_employee_data(employee_data['data'])
                
                if validation_errors:
                    employee_data['errors'] = validation_errors
                    errors.append({
                        'row': index + 1,
                        'errors': validation_errors
                    })
                
                mapped_data.append(employee_data)
            
            return {
                'success': True,
                'data': mapped_data,
                'total_rows': len(mapped_data),
                'error_rows': len(errors),
                'errors': errors
            }
            
        except Exception as e:
            logger.error(f"读取Excel文件失败: {e}")
            return {
                'success': False,
                'error': str(e),
                'data': [],
                'total_rows': 0,
                'error_rows': 0,
                'errors': []
            }
    
    def _map_employee_fields(self, row_data: Dict[str, Any], row_number: int) -> Dict[str, Any]:
        """
        映射员工字段
        
        Args:
            row_data: 行数据
            row_number: 行号
            
        Returns:
            映射后的员工数据
        """
        mapped_data = {
            'row_number': row_number,
            'original_data': row_data,
            'data': {},
            'mapping_info': {},
            'errors': {}
        }
        
        # 遍历字段映射
        for standard_field, possible_names in self.field_mapping.items():
            value = None
            matched_column = None
            
            # 查找匹配的列名
            for col_name in row_data.keys():
                if col_name in possible_names:
                    value = row_data[col_name]
                    matched_column = col_name
                    break
            
            # 如果没有找到精确匹配，尝试模糊匹配
            if value is None:
                for col_name in row_data.keys():
                    for possible_name in possible_names:
                        if possible_name in col_name or col_name in possible_name:
                            value = row_data[col_name]
                            matched_column = col_name
                            break
                    if value is not None:
                        break
            
            # 清理和转换数据
            if value is not None:
                value = self._clean_cell_value(value)
                mapped_data['data'][standard_field] = value
                mapped_data['mapping_info'][standard_field] = {
                    'source_column': matched_column,
                    'original_value': row_data.get(matched_column, ''),
                    'cleaned_value': value
                }
        
        return mapped_data
    
    def _clean_cell_value(self, value: Any) -> str:
        """
        清理单元格值
        
        Args:
            value: 原始值
            
        Returns:
            清理后的字符串值
        """
        if pd.isna(value) or value is None:
            return ''
        
        # 转换为字符串
        str_value = str(value).strip()
        
        # 移除多余的空格
        str_value = re.sub(r'\s+', ' ', str_value)
        
        # 处理数字格式（如身份证号、银行卡号）
        if str_value.endswith('.0'):
            str_value = str_value[:-2]
        
        return str_value
    
    def export_employees_to_excel(self, employees: List[Dict[str, Any]], filename: str = None) -> bytes:
        """
        导出员工数据到Excel
        
        Args:
            employees: 员工数据列表
            filename: 文件名
            
        Returns:
            Excel文件的字节内容
        """
        try:
            # 创建DataFrame
            df_data = []
            for emp in employees:
                row = {
                    '姓名': emp.get('name', ''),
                    '身份证号': emp.get('id_card', ''),
                    '手机号': emp.get('phone', ''),
                    '银行卡号': emp.get('bank_card', ''),
                    '开户行': emp.get('bank_name', ''),
                    '入职日期': emp.get('hire_date', ''),
                    '部门': emp.get('department', ''),
                    '职位': emp.get('position', ''),
                    '状态': '归档' if emp.get('is_archived', False) else '在职',
                    '创建时间': emp.get('created_at', ''),
                    '更新时间': emp.get('updated_at', '')
                }
                df_data.append(row)
            
            df = pd.DataFrame(df_data)
            
            # 创建Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='员工信息', index=False)
                
                # 获取工作表
                worksheet = writer.sheets['员工信息']
                
                # 设置样式
                self._apply_excel_styles(worksheet, df)
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise
    
    def _apply_excel_styles(self, worksheet, df: pd.DataFrame):
        """
        应用Excel样式
        
        Args:
            worksheet: openpyxl工作表
            df: DataFrame数据
        """
        # 标题行样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用标题行样式
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行样式
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_salary_to_excel(self, salary_records: List[Dict[str, Any]], account_info: Dict[str, Any]) -> bytes:
        """
        导出工资表到Excel
        
        Args:
            salary_records: 工资记录列表
            account_info: 账表信息
            
        Returns:
            Excel文件的字节内容
        """
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '工资表'
            
            # 设置标题
            title = f"{account_info.get('name', '工资表')} - {account_info.get('period', '')}"
            ws.merge_cells('A1:H1')
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置表头
            headers = ['序号', '姓名', '基本工资', '奖金', '扣款', '实发工资', '发放日期', '备注']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            
            # 填充数据
            for row, record in enumerate(salary_records, 4):
                ws.cell(row=row, column=1, value=row - 3)
                ws.cell(row=row, column=2, value=record.get('employee_name', ''))
                ws.cell(row=row, column=3, value=record.get('salary_amount', 0))
                ws.cell(row=row, column=4, value=record.get('bonus', 0))
                ws.cell(row=row, column=5, value=record.get('deduction', 0))
                
                # 计算实发工资
                actual_salary = (record.get('salary_amount', 0) + 
                               record.get('bonus', 0) - 
                               record.get('deduction', 0))
                ws.cell(row=row, column=6, value=actual_salary)
                
                ws.cell(row=row, column=7, value=record.get('pay_date', ''))
                ws.cell(row=row, column=8, value=record.get('remarks', ''))
            
            # 添加合计行
            total_row = len(salary_records) + 4
            ws.cell(row=total_row, column=1, value='合计')
            ws.cell(row=total_row, column=3, value=f"=SUM(C4:C{total_row-1})")
            ws.cell(row=total_row, column=4, value=f"=SUM(D4:D{total_row-1})")
            ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{total_row-1})")
            ws.cell(row=total_row, column=6, value=f"=SUM(F4:F{total_row-1})")
            
            # 设置合计行样式
            for col in range(1, 9):
                cell = ws.cell(row=total_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            
            # 应用边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(3, total_row + 1):
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border
            
            # 自动调整列宽
            column_widths = [8, 15, 12, 10, 10, 12, 12, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"导出工资表Excel失败: {e}")
            raise
    
    def export_expense_to_excel(self, expense_records: List[Dict[str, Any]], account_info: Dict[str, Any]) -> bytes:
        """
        导出支出表到Excel
        
        Args:
            expense_records: 支出记录列表
            account_info: 账表信息
            
        Returns:
            Excel文件的字节内容
        """
        try:
            # 创建DataFrame
            df_data = []
            for record in expense_records:
                row = {
                    '项目名称': record.get('project_name', ''),
                    '支出类型': record.get('expense_type', ''),
                    '支出金额': record.get('amount', 0),
                    '支出日期': record.get('expense_date', ''),
                    '备注': record.get('remarks', ''),
                    '创建时间': record.get('created_at', '')
                }
                df_data.append(row)
            
            df = pd.DataFrame(df_data)
            
            # 创建Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='支出明细', index=False)
                
                # 获取工作表
                worksheet = writer.sheets['支出明细']
                
                # 设置样式
                self._apply_excel_styles(worksheet, df)
                
                # 添加合计行
                total_row = len(df) + 2
                worksheet.cell(row=total_row, column=1, value='合计')
                worksheet.cell(row=total_row, column=3, value=f"=SUM(C2:C{len(df)+1})")
                
                # 设置合计行样式
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=total_row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"导出支出表Excel失败: {e}")
            raise
    
    def validate_excel_structure(self, file_content: bytes, expected_type: str = 'employee') -> Dict[str, Any]:
        """
        验证Excel文件结构
        
        Args:
            file_content: Excel文件内容
            expected_type: 期望的文件类型 ('employee', 'salary', 'expense')
            
        Returns:
            验证结果
        """
        try:
            df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
            
            result = {
                'valid': True,
                'errors': [],
                'warnings': [],
                'info': {
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'columns': list(df.columns)
                }
            }
            
            if expected_type == 'employee':
                # 验证员工表结构
                required_fields = ['name']
                found_fields = []
                
                for field, possible_names in self.field_mapping.items():
                    for col in df.columns:
                        if col in possible_names:
                            found_fields.append(field)
                            break
                
                missing_required = [field for field in required_fields if field not in found_fields]
                if missing_required:
                    result['valid'] = False
                    result['errors'].append(f"缺少必需字段: {', '.join(missing_required)}")
                
                if len(found_fields) < 2:
                    result['warnings'].append("识别到的有效字段较少，请检查列名是否正确")
            
            return result
            
        except Exception as e:
            return {
                'valid': False,
                'errors': [f"文件读取失败: {str(e)}"],
                'warnings': [],
                'info': {}
            }
    
    def create_employee_template(self) -> bytes:
        """
        创建员工信息导入模板
        
        Returns:
            Excel模板文件的字节内容
        """
        try:
            # 创建模板数据
            template_data = {
                '姓名': ['张三', '李四'],
                '身份证号': ['110101199001011234', '110101199002022345'],
                '手机号': ['13800138000', '13900139000'],
                '银行卡号': ['6222021234567890123', '6222021234567890124'],
                '开户行': ['中国工商银行北京分行', '中国建设银行北京分行'],
                '部门': ['技术部', '销售部'],
                '职位': ['工程师', '销售经理']
            }
            
            df = pd.DataFrame(template_data)
            
            # 创建Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='员工信息模板', index=False)
                
                # 获取工作表
                worksheet = writer.sheets['员工信息模板']
                
                # 应用样式
                self._apply_excel_styles(worksheet, df)
                
                # 添加说明
                worksheet.cell(row=len(df) + 3, column=1, value='说明：')
                worksheet.cell(row=len(df) + 4, column=1, value='1. 姓名、身份证号、手机号为必填项')
                worksheet.cell(row=len(df) + 5, column=1, value='2. 请确保身份证号、手机号格式正确')
                worksheet.cell(row=len(df) + 6, column=1, value='3. 银行卡号和开户行可选填')
                worksheet.cell(row=len(df) + 7, column=1, value='4. 请删除示例数据后填入真实数据')
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"创建模板失败: {e}")
            raise

def create_excel_handler() -> ExcelHandler:
    """
    创建Excel处理器实例
    
    Returns:
        ExcelHandler实例
    """
    return ExcelHandler()